#!/usr/bin/env python3
"""
Create a sample Excel file for testing the Excel Image Reviewer.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def create_sample_excel(filename="sample_data.xlsx"):
    """Create a sample Excel file with multiple sheets."""
    
    # Create a workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Employee Data
    ws1 = wb.create_sheet("Employee Data")
    ws1.append(["Employee ID", "Name", "Department", "Salary", "Start Date"])
    
    # Add some data with intentional issues
    data1 = [
        [1, "John Doe", "Enginering", 75000, "2022-01-15"],  # Misspelling: Enginering
        [2, "Jane Smith", "Marketing", 65000, "2022-03-20"],
        [3, "Bob Jonson", "Sales", 60000, "2023-05-10"],  # Misspelling: Jonson
        [4, "Alice Brown", "Engineering", 80000, "2021-12-01"],
        [5, "Charlie Davis", "Marketing", None, "2022-07-15"],  # Missing salary
    ]
    
    for row in data1:
        ws1.append(row)
    
    # Sheet 2: Sales Data
    ws2 = wb.create_sheet("Sales Data")
    ws2.append(["Date", "Product", "Quantity", "Revenue", "Region"])
    
    # Add data with logical issues
    data2 = [
        ["2023-01-15", "Widget A", 100, 5000, "North"],
        ["2023-02-20", "Widget B", -5, 250, "South"],  # Negative quantity
        ["2023-03-10", "Widget A", 150, 7500, "East"],
        ["2022-12-30", "Widget C", 200, 10000, "West"],  # Date before others
        ["2023-04-05", "Widget B", 80, 4000, "North"],
    ]
    
    for row in data2:
        ws2.append(row)
    
    # Sheet 3: Budget Summary
    ws3 = wb.create_sheet("Budget Summary")
    ws3.append(["Category", "Q1", "Q2", "Q3", "Q4", "Total"])
    
    data3 = [
        ["Marketing", 50000, 60000, 55000, 65000, 230000],
        ["Development", 100000, 110000, 105000, 115000, 430000],
        ["Sales", 75000, 80000, 78000, 82000, 315000],
        ["Operations", 40000, 45000, 42000, 48000, 175000],
        ["Total", 265000, 295000, 280000, 310000, 1150001],  # Math error in total
    ]
    
    for row in data3:
        ws3.append(row)
    
    # Save the workbook
    wb.save(filename)
    print(f"Sample Excel file created: {filename}")
    print("This file contains intentional errors for testing:")
    print("  - Misspellings: 'Enginering', 'Jonson'")
    print("  - Missing data: Salary for Charlie Davis")
    print("  - Logical issues: Negative quantity, date inconsistency")
    print("  - Math error: Incorrect total in Budget Summary")


if __name__ == "__main__":
    create_sample_excel()
